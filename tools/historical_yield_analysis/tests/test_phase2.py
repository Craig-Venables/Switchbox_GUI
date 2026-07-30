"""Phase 2 tests: fabrication join, missing Excel, Origin TXT."""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from historical_yield.analysis import filter_samples, sample_dataframe
from historical_yield.cache import YieldCache
from historical_yield.config import AppConfig, DataRoot
from historical_yield.fabrication import load_fabrication_index
from historical_yield.import_pipeline import scan_and_update_cache
from historical_yield.missing_excel import (
    find_missing_excel,
    folder_has_classification_excel,
    is_dxx_sample_folder,
)
from historical_yield.origin_export import export_origin_txt


def _write_xlsx(path: Path, rows):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for r_i, row in enumerate(rows, start=1):
        for c_i, val in enumerate(row, start=1):
            ws.cell(r_i, c_i, val)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_fab_workbook(path: Path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Memristor Devices"
    headers = [
        "#",
        "Device Full Name",
        "Short Name",
        "B-Material",
        "T-Material",
        "Np Type",
        "Np Concentraion",
        "Polymer",
        "Layer 1",
        "Date Made",
    ]
    ws.append(headers)
    ws.append(
        [
            95,
            "D95-0.1mgml-ITO-PMMA(2%)-Gold-s3",
            "short",
            "ITO",
            "Gold",
            "Zn-Cu-In-S(ZnS)",
            0.1,
            "PMMA",
            "PMMA(2%)",
            "2024-01-01",
        ]
    )
    ws.append(
        [
            10,
            "D10-0.2mgml-Gold-PMMA(2%)-Gold-s3",
            "short10",
            "Gold",
            "Gold",
            "Zn-Cu-In-S(ZnS)",
            0.2,
            "PMMA",
            "PMMA(2%)",
            None,
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def test_is_dxx_folder_rules():
    assert is_dxx_sample_folder("D95-0.1mgml-ITO-PMMA(2%)-Gold-s3")
    assert is_dxx_sample_folder("d10-stock")
    assert not is_dxx_sample_folder("Quantum Dots")
    assert not is_dxx_sample_folder("device_status")
    assert not is_dxx_sample_folder("Sample95")


def test_fabrication_lookup_by_number_and_name(tmp_path: Path):
    fab_path = tmp_path / "solutions.xlsx"
    _write_fab_workbook(fab_path)
    idx = load_fabrication_index(fab_path, "Memristor Devices")
    assert idx.n_rows == 2
    by_num = idx.lookup(sample_number=95)
    assert by_num is not None
    assert by_num.sample_id == "D95"
    assert by_num.polymer == "PMMA"
    assert by_num.polymer_percent == pytest.approx(2.0)
    assert by_num.bottom_electrode == "ITO"
    assert by_num.top_electrode == "Gold"
    assert by_num.concentration_mgml == pytest.approx(0.1)
    by_id = idx.lookup(sample_id="D10")
    assert by_id is not None
    assert by_id.np_type.startswith("Zn-Cu")
    by_name = idx.lookup(device_full_name="D95-0.1mgml-ITO-PMMA(2%)-Gold-s3")
    assert by_name is not None
    assert by_name.sample_number == 95


def test_fabrication_missing_file_soft_fails(tmp_path: Path):
    idx = load_fabrication_index(tmp_path / "nope.xlsx", "Memristor Devices")
    assert idx.n_rows == 0
    assert idx.load_error
    assert "not found" in idx.load_error.lower()


def test_fabrication_permission_error_soft_fails(tmp_path: Path, monkeypatch):
    fab_path = tmp_path / "solutions.xlsx"
    _write_fab_workbook(fab_path)

    def boom(*_args, **_kwargs):
        raise PermissionError("Permission denied")

    monkeypatch.setattr(
        "historical_yield.fabrication._open_workbook",
        boom,
    )
    idx = load_fabrication_index(fab_path, "Memristor Devices")
    assert idx.n_rows == 0
    assert idx.load_error
    assert "Permission denied" in idx.load_error


def test_missing_excel_detection(tmp_path: Path):
    root = tmp_path / "data"
    # has excel
    has = root / "D10-0.2mgml-Gold-PMMA(2%)-Gold-s3"
    has.mkdir(parents=True)
    _write_xlsx(has / f"{has.name}.xlsx", [["Section", "Device #", "Classification"], ["A", 1, "Ohmic"]])
    # missing excel
    missing = root / "D11-Stock-Gold-PS(2%)-Gold-s4"
    missing.mkdir(parents=True)
    # non-dxx ignored
    other = root / "OtherProject"
    other.mkdir()
    (other / "notes.xlsx").write_text("x", encoding="utf-8")

    assert folder_has_classification_excel(has, "D10")
    assert not folder_has_classification_excel(missing, "D11")

    config = AppConfig(
        data_roots=[DataRoot(name="data", path=root, priority=1)],
        cache_dir=tmp_path / "cache",
        output_dir=tmp_path / "output",
    )
    entries = find_missing_excel(config)
    ids = {e.sample_id for e in entries}
    assert "D11" in ids
    assert "D10" not in ids
    assert all(e.sample_id.startswith("D") for e in entries)


def test_sample_dataframe_joins_fab(tmp_path: Path):
    root = tmp_path / "data"
    sample_dir = root / "D10-0.2mgml-Gold-PMMA(2%)-Gold-s3"
    xlsx = sample_dir / f"{sample_dir.name}.xlsx"
    _write_xlsx(
        xlsx,
        [
            ["Section", "Device #", "Classification"],
            ["A", 1, "Memristive"],
            ["A", 2, "Ohmic"],
        ],
    )
    fab_path = tmp_path / "solutions.xlsx"
    _write_fab_workbook(fab_path)
    config = AppConfig(
        data_roots=[DataRoot(name="data", path=root, priority=1)],
        cache_dir=tmp_path / "cache",
        output_dir=tmp_path / "output",
        fabrication_workbook=fab_path,
    )
    scan_and_update_cache(config, rebuild=True)
    cache = YieldCache(config.sqlite_path)
    df = sample_dataframe(
        cache,
        fabrication_workbook=fab_path,
        fabrication_sheet="Memristor Devices",
    )
    assert len(df) == 1
    assert bool(df.iloc[0]["has_fab_row"]) is True
    assert df.iloc[0]["polymer"] == "PMMA"
    assert df.iloc[0]["polymer_percent"] == pytest.approx(2.0)
    assert df.iloc[0]["np_type"] == "Zn-Cu-In-S(ZnS)"


def test_origin_txt_tab_separated(tmp_path: Path):
    import pandas as pd

    df = pd.DataFrame(
        [
            {
                "sample_id": "D10",
                "sample_number": 10,
                "sample_name": "D10-x",
                "strict_yield": 0.5,
                "strict_yield_pct": 50.0,
                "n_classified": 2,
                "n_memristive": 1,
                "n_ohmic": 1,
                "n_capacitive": 0,
                "n_conductive": 0,
                "n_non_conductive": 0,
                "n_mem_capacitive": 0,
                "n_intermittent": 0,
                "n_other": 0,
                "pct_memristive": 50.0,
                "pct_ohmic": 50.0,
                "pct_capacitive": 0.0,
                "pct_conductive": 0.0,
                "pct_non_conductive": 0.0,
                "pct_mem_capacitive": 0.0,
                "pct_intermittent": 0.0,
                "pct_other": 0.0,
                "polymer": "PMMA",
                "polymer_percent": 2.0,
                "bottom_electrode": "Gold",
                "top_electrode": "Gold",
                "np_type": "QD",
                "concentration_mgml": 0.2,
                "is_stock": False,
            }
        ]
    )
    paths = export_origin_txt(df, tmp_path / "origin")
    assert len(paths) == 3
    text = paths[0].read_text(encoding="utf-8")
    assert "\t" in text.splitlines()[0]
    assert "sample_id" in text.splitlines()[0]
    assert "," not in text.splitlines()[0] or "sample_id" in text  # header is TSV
    assert text.splitlines()[1].startswith("D10\t")


def test_filter_selection_reduces_rows():
    import pandas as pd

    df = pd.DataFrame(
        [
            {
                "sample_id": "D1",
                "sample_number": 1,
                "polymer": "PMMA",
                "bottom_electrode": "ITO",
                "top_electrode": "Gold",
                "polymer_percent": 2.0,
                "np_type": "A",
            },
            {
                "sample_id": "D2",
                "sample_number": 2,
                "polymer": "PS",
                "bottom_electrode": "Gold",
                "top_electrode": "Gold",
                "polymer_percent": 2.0,
                "np_type": "A",
            },
        ]
    )
    filtered = filter_samples(df, polymers=["PMMA"])
    assert list(filtered["sample_id"]) == ["D1"]
    filtered2 = filter_samples(df, sample_ids=["D2"], bottom_electrodes=["Gold"])
    assert list(filtered2["sample_id"]) == ["D2"]
