"""Load fabrication metadata from solutions and devices.xlsx (read-only)."""

from __future__ import annotations

import re
import shutil
import tempfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Dict, Optional, Tuple

from .parse_sample import extract_sample_id, parse_concentration


@dataclass
class FabricationRecord:
    sample_number: int
    sample_id: str
    device_full_name: Optional[str] = None
    short_name: Optional[str] = None
    polymer: Optional[str] = None
    polymer_percent: Optional[float] = None
    concentration_raw: Optional[str] = None
    concentration_mgml: Optional[float] = None
    is_stock: bool = False
    bottom_electrode: Optional[str] = None
    top_electrode: Optional[str] = None
    np_type: Optional[str] = None
    date_made: Optional[str] = None
    layer_1: Optional[str] = None
    annealing: Optional[str] = None
    control: Optional[str] = None
    extra: Dict[str, Any] | None = None

    def to_dict(self) -> Dict[str, Any]:
        d = asdict(self)
        return d


_POLYMER_PCT_RE = re.compile(r"(?i)\((\d+(?:\.\d+)?)\s*%\)")


def _cell(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _polymer_percent_from_layer(layer: Optional[str], polymer: Optional[str]) -> Optional[float]:
    for source in (layer, polymer):
        if not source:
            continue
        m = _POLYMER_PCT_RE.search(str(source))
        if m:
            return float(m.group(1))
    return None


def _concentration_from_fab(raw: Any) -> tuple[Optional[str], Optional[float], bool]:
    if raw is None:
        return None, None, False
    if isinstance(raw, (int, float)):
        return f"{float(raw)}mgml", float(raw), False
    text = str(raw).strip()
    if not text:
        return None, None, False
    return parse_concentration(text)


class FabricationIndex:
    """In-memory index of Memristor Devices rows keyed by D# and full name."""

    def __init__(self):
        self.by_number: Dict[int, FabricationRecord] = {}
        self.by_sample_id: Dict[str, FabricationRecord] = {}
        self.by_full_name: Dict[str, FabricationRecord] = {}
        self.source_path: Optional[Path] = None
        self.sheet_name: Optional[str] = None
        self.n_rows: int = 0
        self.load_error: Optional[str] = None

    def lookup(
        self,
        *,
        sample_number: Optional[int] = None,
        sample_id: Optional[str] = None,
        device_full_name: Optional[str] = None,
    ) -> Optional[FabricationRecord]:
        if sample_id and sample_id in self.by_sample_id:
            return self.by_sample_id[sample_id]
        if sample_number is not None and sample_number in self.by_number:
            return self.by_number[sample_number]
        if device_full_name:
            key = device_full_name.strip()
            if key in self.by_full_name:
                return self.by_full_name[key]
            sid = extract_sample_id(key)
            if sid and sid[0] in self.by_sample_id:
                return self.by_sample_id[sid[0]]
        if sample_id:
            sid = extract_sample_id(sample_id)
            if sid and sid[1] in self.by_number:
                return self.by_number[sid[1]]
        return None

    def has_sample(self, sample_id: str) -> bool:
        return self.lookup(sample_id=sample_id) is not None


def _open_workbook(path: Path):
    """Open an xlsx, falling back to a temp copy if the source is locked."""
    import openpyxl

    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True), None
    except PermissionError:
        # Excel / OneDrive often exclusive-locks the live file; a copy usually works.
        tmp = Path(tempfile.gettempdir()) / f"historical_yield_fab_copy_{path.stem}.xlsx"
        try:
            shutil.copy2(path, tmp)
            return openpyxl.load_workbook(tmp, read_only=True, data_only=True), tmp
        except Exception as copy_exc:
            raise PermissionError(
                f"Permission denied opening {path}. Close it in Excel if open, "
                f"then retry. Also failed to read via temp copy: {copy_exc}"
            ) from copy_exc


def load_fabrication_index(
    workbook_path: Path | str | None,
    sheet_name: str = "Memristor Devices",
) -> FabricationIndex:
    """Load the fabrication workbook. Returns empty index if path missing/unreadable."""
    index = FabricationIndex()
    if not workbook_path:
        return index
    path = Path(workbook_path)
    index.source_path = path
    index.sheet_name = sheet_name
    if not path.exists():
        index.load_error = f"file not found: {path}"
        return index

    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:  # pragma: no cover
        raise ImportError("openpyxl is required to read fabrication workbook") from exc

    tmp_copy: Optional[Path] = None
    try:
        wb, tmp_copy = _open_workbook(path)
    except Exception as exc:
        index.load_error = str(exc)
        return index

    try:
        if sheet_name not in wb.sheetnames:
            index.load_error = f"sheet not found: {sheet_name!r}"
            return index
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return index
        cols = {
            str(h).strip(): i
            for i, h in enumerate(header)
            if h is not None and str(h).strip()
        }

        def get(row, *names):
            for name in names:
                if name in cols and cols[name] < len(row):
                    return row[cols[name]]
            return None

        for row in rows:
            if not row:
                continue
            num_raw = get(row, "#")
            full_name = _cell(get(row, "Device Full Name"))
            sample_number: Optional[int] = None
            sample_id: Optional[str] = None
            if num_raw is not None and str(num_raw).strip() != "":
                try:
                    sample_number = int(float(str(num_raw).strip()))
                    sample_id = f"D{sample_number}"
                except (TypeError, ValueError):
                    sample_number = None
            if sample_id is None and full_name:
                sid = extract_sample_id(full_name)
                if sid:
                    sample_id, sample_number = sid
            if sample_id is None or sample_number is None:
                continue

            conc_raw, conc_mgml, is_stock = _concentration_from_fab(
                get(row, "Np Concentraion", "Np Concentration")
            )
            polymer = _cell(get(row, "Polymer"))
            layer_1 = _cell(get(row, "Layer 1"))
            polymer_percent = _polymer_percent_from_layer(layer_1, polymer)
            # Prefer polymer base name without percent decoration when possible
            polymer_name = polymer
            if polymer_name and "(" in polymer_name:
                polymer_name = polymer_name.split("(", 1)[0].strip() or polymer

            rec = FabricationRecord(
                sample_number=sample_number,
                sample_id=sample_id,
                device_full_name=full_name,
                short_name=_cell(get(row, "Short Name")),
                polymer=polymer_name,
                polymer_percent=polymer_percent,
                concentration_raw=conc_raw,
                concentration_mgml=conc_mgml,
                is_stock=is_stock,
                bottom_electrode=_cell(get(row, "B-Material")),
                top_electrode=_cell(get(row, "T-Material")),
                np_type=_cell(get(row, "Np Type")),
                date_made=_cell(get(row, "Date Made")),
                layer_1=layer_1,
                annealing=_cell(get(row, "Annealing")),
                control=_cell(get(row, "Controll?", "Control?")),
            )
            index.by_number[sample_number] = rec
            index.by_sample_id[sample_id] = rec
            if full_name:
                index.by_full_name[full_name] = rec
            index.n_rows += 1
    except Exception as exc:
        index.load_error = str(exc)
        return index
    finally:
        try:
            wb.close()
        except Exception:
            pass
        if tmp_copy is not None:
            try:
                tmp_copy.unlink(missing_ok=True)
            except OSError:
                pass
    return index


# Success-only cache so a locked-file failure can be retried after Excel is closed.
_FAB_CACHE: Dict[Tuple[str, str], FabricationIndex] = {}


def clear_fabrication_cache() -> None:
    _FAB_CACHE.clear()


def cached_fabrication_index(workbook_path: str, sheet_name: str) -> FabricationIndex:
    """Backward-compatible helper; only caches successful loads."""
    return get_fabrication_index(workbook_path, sheet_name, use_cache=True)


def get_fabrication_index(
    workbook_path: Path | str | None,
    sheet_name: str = "Memristor Devices",
    *,
    use_cache: bool = True,
) -> FabricationIndex:
    if not workbook_path:
        return FabricationIndex()
    path = str(Path(workbook_path).resolve())
    key = (path, sheet_name)
    if use_cache and key in _FAB_CACHE:
        return _FAB_CACHE[key]
    index = load_fabrication_index(path, sheet_name)
    if use_cache and not index.load_error:
        _FAB_CACHE[key] = index
    return index


def enrich_sample_row(row: Dict[str, Any], fab: Optional[FabricationRecord]) -> Dict[str, Any]:
    """Overlay fabrication fields onto a sample summary dict (fab wins when present)."""
    out = dict(row)
    out["has_fab_row"] = fab is not None
    if fab is None:
        out.setdefault("polymer_percent", None)
        out.setdefault("np_type", None)
        out.setdefault("date_made", None)
        out.setdefault("device_full_name", None)
        return out

    out["device_full_name"] = fab.device_full_name
    out["np_type"] = fab.np_type
    out["date_made"] = fab.date_made
    out["polymer_percent"] = fab.polymer_percent
    if fab.polymer:
        out["polymer"] = fab.polymer
    if fab.bottom_electrode:
        out["bottom_electrode"] = fab.bottom_electrode
    if fab.top_electrode:
        out["top_electrode"] = fab.top_electrode
    if fab.concentration_mgml is not None:
        out["concentration_mgml"] = fab.concentration_mgml
        out["is_stock"] = fab.is_stock
    elif fab.is_stock:
        out["concentration_mgml"] = 0.0
        out["is_stock"] = True
    return out
