"""Read and parse per-sample classification Excel workbooks."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from .models import DeviceClassification
from .normalize import is_instruction_label, normalize_classification


# Map header synonyms -> canonical field names used on DeviceClassification
_HEADER_ALIASES = {
    "section": "section",
    "section ": "section",
    "device #": "device_number",
    "device number": "device_number",
    "device#": "device_number",
    "classification": "classification",
    "memristor strength": "memristor_strength",
    "how strong": "memristor_strength",
    "current range": "current_range",
    "resistance value": "resistance_value",
    "# sweeps": "n_sweeps",
    "number of sweeps": "n_sweeps",
    "retention": "retention",
    "retention ": "retention",
    "endurance": "endurance",
    "volatile": "volatile",
    "current state": "current_state",
    "date measured": "date_measured",
    "notes": "notes",
    "forming step": "forming_step",
    "total yield": "total_yield",
    "devices measured": "devices_measured",
    "working memristors": "working_memristors",
}


def _cell_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _looks_like_header(row: Sequence[Any]) -> bool:
    lowered = [str(c).strip().lower() if c is not None else "" for c in row]
    return "classification" in lowered and any(
        x in lowered for x in ("section", "section ", "device #", "device number")
    )


def _looks_like_instruction_row(row: Sequence[Any]) -> bool:
    texts = [_cell_str(c) for c in row]
    joined = " ".join(t for t in texts if t).lower()
    if "device section" in joined or "device number" in joined:
        return True
    if any(is_instruction_label(t) for t in texts if t):
        return True
    # second template row often has long explanatory phrases
    longish = sum(1 for t in texts if t and len(t) > 25)
    return longish >= 2


def map_headers(header_row: Sequence[Any]) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        field = _HEADER_ALIASES.get(key)
        if field:
            mapping[idx] = field
    return mapping


def parse_workbook_rows(
    rows: Iterable[Sequence[Any]],
    *,
    success_categories: Optional[Iterable[str]] = None,
) -> Tuple[List[DeviceClassification], List[str], Optional[str]]:
    """
    Parse an already-loaded sheet (iterable of row tuples).

    Returns (devices, warnings, schema_header_string).
    """
    warnings: List[str] = []
    row_list = [tuple(r) for r in rows]
    if not row_list:
        return [], ["empty workbook"], None

    header_idx = None
    for i, row in enumerate(row_list[:5]):
        if _looks_like_header(row):
            header_idx = i
            break
    if header_idx is None:
        # fallback: first non-empty row
        for i, row in enumerate(row_list[:3]):
            if any(c is not None and str(c).strip() for c in row):
                header_idx = i
                warnings.append("classification header not found; using first non-empty row")
                break
    if header_idx is None:
        return [], ["no header row found"], None

    header = row_list[header_idx]
    colmap = map_headers(header)
    schema_header = " | ".join(
        str(c).strip() if c is not None else "" for c in header
    ).strip(" |")

    if "classification" not in colmap.values():
        warnings.append("no Classification column mapped")
    if "section" not in colmap.values() or "device_number" not in colmap.values():
        warnings.append("missing Section and/or Device # columns")

    devices: List[DeviceClassification] = []
    data_start = header_idx + 1
    # skip instruction row if present
    if data_start < len(row_list) and _looks_like_instruction_row(row_list[data_start]):
        data_start += 1

    for row in row_list[data_start:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        fields: Dict[str, Any] = {}
        for idx, name in colmap.items():
            if idx < len(row):
                fields[name] = row[idx]

        section = _cell_str(fields.get("section"))
        if section is None:
            continue
        # skip leftover header-like section labels
        if section.lower().startswith("device section"):
            continue
        section = section.strip().upper()[:1] if len(section.strip()) == 1 else section.strip()

        raw_num = fields.get("device_number")
        try:
            if raw_num is None or str(raw_num).strip() == "":
                continue
            device_number = int(float(str(raw_num).strip()))
        except (TypeError, ValueError):
            warnings.append(f"skipping row with bad device number: {raw_num!r}")
            continue

        raw_cls = _cell_str(fields.get("classification"))
        canonical, is_classified, is_success = normalize_classification(
            raw_cls, success_categories=success_categories
        )
        if raw_cls and not is_classified and is_instruction_label(raw_cls):
            continue

        extra = {}
        for k in ("forming_step", "total_yield", "devices_measured", "working_memristors"):
            if k in fields and fields[k] is not None:
                extra[k] = _cell_str(fields[k])

        devices.append(
            DeviceClassification(
                section=section,
                device_number=device_number,
                raw_classification=raw_cls,
                normalized_classification=canonical,
                is_classified=is_classified,
                is_yield_success=is_success,
                memristor_strength=_cell_str(fields.get("memristor_strength")),
                current_range=_cell_str(fields.get("current_range")),
                resistance_value=_cell_str(fields.get("resistance_value")),
                n_sweeps=_cell_str(fields.get("n_sweeps")),
                retention=_cell_str(fields.get("retention")),
                endurance=_cell_str(fields.get("endurance")),
                volatile=_cell_str(fields.get("volatile")),
                current_state=_cell_str(fields.get("current_state")),
                date_measured=_cell_str(fields.get("date_measured")),
                notes=_cell_str(fields.get("notes")),
                extra=extra,
            )
        )

    if not devices:
        warnings.append("no device rows parsed")
    return devices, warnings, schema_header or None


def load_workbook_devices(
    path: Path | str,
    *,
    success_categories: Optional[Iterable[str]] = None,
) -> Tuple[List[DeviceClassification], List[str], Optional[str]]:
    """Open an xlsx and parse the first sheet."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise ImportError("openpyxl is required to read classification workbooks") from exc

    path = Path(path)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"cannot open workbook: {exc}") from exc

    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    return parse_workbook_rows(rows, success_categories=success_categories)
